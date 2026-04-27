"""taiji-runner — per-sample Taiji 1.3.0 orchestrator.

Bridges the gap between our xlsx (the human-readable manifest produced by
build-taiji-input) and Taiji's actual input format (one TSV + one config
per sample, since `taiji run` operates on a single sample at a time).

Mirrors the architecture of the existing UCSD wrapper at
Taiji/taiji_wrapper-uniqueGen.py — same directory layout, same three-
placeholder config substitution — but adds:
  - Sequential or parallel local execution (no SLURM dependency)
  - Per-sample workflow-log integration
  - Path-with-spaces safety
  - Output validation (Taiji exits 0 on workflow failure; we cross-check
    that pagerank files actually got produced before reporting success)

Layout produced under <run_dir>/:

    Input/
      <sample>_input/
        <sample>_input.tsv      (filtered Active sheet)
        <sample>_config.yml     (template with 3 placeholders filled)
    Output/
      Partial/
        <sample>_output/        (taiji's cwd; GeneRanks.tsv etc. land here)
    taiji_config_files.txt      (newline-separated list of per-sample configs)

CLI
---

    python run_taiji.py \\
        --xlsx       runs/<name>/taiji_input.xlsx \\
        --template   runs/<name>/taiji_config.template.yml \\
        --binary     binaries/taiji-macOS-Catalina-10.15 \\
        --run-dir    runs/<name> \\
        [--repo-root <abs path>]   # for ${REPO_ROOT} substitution; default = autodetect
        [--threads 4]              # passed as +RTS -N<n>
        [--parallel 1]             # how many samples to run concurrently (default 1)
        [--prepare-only]           # write per-sample TSVs + configs; do NOT invoke taiji
        [--samples RA_11,OA_02]    # restrict to a subset of samples
        [--continue-on-error]      # keep going if one sample fails
"""

from __future__ import annotations

import argparse
import concurrent.futures
import json
import os
import shutil
import subprocess
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

try:
    import openpyxl
except ImportError:
    print("taiji-runner: openpyxl required (`micromamba install openpyxl`)",
          file=sys.stderr)
    sys.exit(2)


SCRIPT_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = SCRIPT_DIR.parent / "templates"
SBATCH_TEMPLATE = TEMPLATE_DIR / "taiji_array.sbatch.template"


# ---------------------------------------------------------------------------
# Soft-attach to workflow-log
# ---------------------------------------------------------------------------


def _attach_log(work_dir: Path | None = None):
    """Soft-attach to the workflow-log skill. `work_dir` should be the run
    directory where `log/active_run` lives — NOT cwd, which may differ
    when this skill is called from a parent shell wrapper that doesn't
    cd into the run dir first."""
    try:
        sibling = SCRIPT_DIR.parent.parent / "workflow-log" / "scripts"
        if sibling.exists() and str(sibling) not in sys.path:
            sys.path.insert(0, str(sibling))
        from log import TaijiLog  # type: ignore[import-not-found]
        return TaijiLog.attach_active(work_dir=work_dir or Path.cwd())
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------


@dataclass
class SampleSpec:
    name: str             # e.g. "RA_11"
    rows: list[dict]      # list of Active-sheet dicts (one per modality row)
    genome_fasta: str     # path from active_metadata.vcf_Location
    gtf_path: str | None  # active_metadata.gtf_Location if useful for sanity


@dataclass
class SampleResult:
    sample: str
    config_path: Path
    tsv_path: Path
    output_dir: Path
    exit_code: int = -1
    duration_s: float = 0.0
    n_generanks: int = 0       # GeneRanks.tsv (the headline PageRank output)
    n_edges: int = 0           # Network/<sample>/edges_combined.csv
    n_nodes: int = 0           # Network/<sample>/nodes.csv
    stderr_tail: str = ""
    stdout_log: Path | None = None
    stderr_log: Path | None = None
    error: str | None = None

    @property
    def ok(self) -> bool:
        # Taiji 1.3.0 sometimes exits 0 even when its workflow errored; require
        # at least one GeneRanks.tsv (the PageRank-equivalent output) as proof
        # of a real success.
        return self.exit_code == 0 and self.n_generanks > 0 and self.error is None


# ---------------------------------------------------------------------------
# xlsx -> per-sample split
# ---------------------------------------------------------------------------


def _xlsx_sheet_to_dicts(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [h for h in rows[0] if h is not None]
    out = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        out.append({h: row[i] if i < len(row) else None
                    for i, h in enumerate(headers)})
    return out


def split_xlsx(xlsx_path: Path) -> list[SampleSpec]:
    """Read both sheets; return one SampleSpec per Submitter_ID in active_metadata."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Active" not in wb.sheetnames or "active_metadata" not in wb.sheetnames:
        raise SystemExit(
            f"xlsx {xlsx_path} missing required sheets "
            "(need both 'Active' and 'active_metadata')."
        )
    active_rows = _xlsx_sheet_to_dicts(wb["Active"])
    md_rows = _xlsx_sheet_to_dicts(wb["active_metadata"])

    samples: list[SampleSpec] = []
    for md in md_rows:
        sample = md.get("Submitter_ID")
        if not sample:
            continue
        rows = [r for r in active_rows if str(r.get("group", "")) == str(sample)]
        if not rows:
            print(f"WARN: no Active rows for sample '{sample}' — skipping",
                  file=sys.stderr)
            continue
        samples.append(SampleSpec(
            name=str(sample),
            rows=rows,
            genome_fasta=str(md.get("vcf_Location") or ""),
            gtf_path=str(md.get("gtf_Location") or "") or None,
        ))
    return samples


# ---------------------------------------------------------------------------
# Per-sample materialization
# ---------------------------------------------------------------------------


ACTIVE_COLS = ("type", "id", "group", "rep", "path", "tags", "format", "cohort")


def write_per_sample_tsv(sample: SampleSpec, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w") as fh:
        fh.write("\t".join(ACTIVE_COLS) + "\n")
        for row in sample.rows:
            fh.write("\t".join(
                "" if row.get(c) is None else str(row.get(c))
                for c in ACTIVE_COLS) + "\n")


def materialize_config(template_path: Path, out_path: Path, *,
                       input_filepath: Path, output_directory: Path,
                       genome_filepath: str, repo_root: Path) -> None:
    """Substitute the three formula placeholders + ${REPO_ROOT} into the
    template; write to out_path."""
    text = template_path.read_text()
    text = text.replace("[insert_input_filepath_here]", str(input_filepath))
    text = text.replace("[insert_output_directory_here]", str(output_directory))
    text = text.replace("[insert_genome_filepath_here]", genome_filepath)
    text = text.replace("${REPO_ROOT}", str(repo_root))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(text)


def prepare_sample(sample: SampleSpec, run_dir: Path,
                   template_path: Path, repo_root: Path) -> SampleResult:
    sample_input_dir = run_dir / "Input" / f"{sample.name}_input"
    sample_output_dir = run_dir / "Output" / "Partial" / f"{sample.name}_output"
    sample_input_dir.mkdir(parents=True, exist_ok=True)
    sample_output_dir.mkdir(parents=True, exist_ok=True)

    tsv_path = sample_input_dir / f"{sample.name}_input.tsv"
    config_path = sample_input_dir / f"{sample.name}_config.yml"

    write_per_sample_tsv(sample, tsv_path)
    materialize_config(
        template_path, config_path,
        input_filepath=tsv_path,
        output_directory=sample_output_dir,
        genome_filepath=sample.genome_fasta,
        repo_root=repo_root,
    )
    return SampleResult(
        sample=sample.name,
        config_path=config_path,
        tsv_path=tsv_path,
        output_dir=sample_output_dir,
    )


# ---------------------------------------------------------------------------
# Per-sample invocation
# ---------------------------------------------------------------------------


def _count_outputs(output_dir: Path) -> tuple[int, int, int]:
    """Count Taiji 1.3.0 output artifacts: GeneRanks.tsv (the headline output),
    edges_combined.csv (TF -> target edges, under Network/<sample>/), and
    nodes.csv (graph nodes). Returns (n_generanks, n_edges, n_nodes)."""
    if not output_dir.exists():
        return 0, 0, 0
    # Taiji 1.3.0 writes GeneRanks.tsv at the top level of output_dir; in
    # multi-sample setups rglob handles either layout.
    n_generanks = sum(1 for _ in output_dir.rglob("GeneRanks.tsv"))
    n_edges     = sum(1 for _ in output_dir.rglob("edges_combined.csv"))
    n_nodes     = sum(1 for _ in output_dir.rglob("nodes.csv"))
    return n_generanks, n_edges, n_nodes


def invoke_taiji(result: SampleResult, binary: Path, threads: int,
                 log_dir: Path) -> SampleResult:
    """Run `taiji run --config <result.config_path> +RTS -N<threads>` from
    inside result.output_dir. Capture stdout/stderr to log_dir/."""
    log_dir.mkdir(parents=True, exist_ok=True)
    result.stdout_log = log_dir / f"{result.sample}.taiji.stdout"
    result.stderr_log = log_dir / f"{result.sample}.taiji.stderr"

    cmd = [str(binary), "run", "--config", str(result.config_path),
           "+RTS", f"-N{threads}"]
    print(f"[taiji-runner] [{result.sample}] running: {' '.join(cmd)}",
          file=sys.stderr)
    print(f"[taiji-runner] [{result.sample}]   cwd={result.output_dir}",
          file=sys.stderr)

    t0 = time.time()
    try:
        with result.stdout_log.open("w") as out_fh, \
             result.stderr_log.open("w") as err_fh:
            proc = subprocess.run(cmd, cwd=str(result.output_dir),
                                  stdout=out_fh, stderr=err_fh, check=False)
        result.exit_code = proc.returncode
    except Exception as e:
        result.error = f"subprocess raised: {e}"
        result.exit_code = -1
    result.duration_s = round(time.time() - t0, 1)

    # Cross-check: even if exit_code == 0, Taiji may have silently failed.
    result.n_generanks, result.n_edges, result.n_nodes = _count_outputs(result.output_dir)

    if result.stderr_log and result.stderr_log.exists():
        try:
            with result.stderr_log.open() as fh:
                lines = fh.readlines()
                result.stderr_tail = "".join(lines[-20:])[-2000:]
        except OSError:
            pass

    if result.exit_code == 0 and result.n_generanks == 0:
        result.error = ("Taiji exited 0 but produced no GeneRanks.tsv — "
                        "silent workflow failure. Inspect stderr.")
    return result


# ---------------------------------------------------------------------------
# Executor: local (current behavior) vs SLURM (array job)
# ---------------------------------------------------------------------------


def detect_executor(user_choice: str) -> str:
    """Resolve --executor=auto into 'slurm' or 'local'. 'auto' picks SLURM
    when `sbatch` is on PATH, else local. Explicit choices are honored
    strictly (no fallback)."""
    if user_choice in ("local", "slurm"):
        return user_choice
    return "slurm" if shutil.which("sbatch") else "local"


def _slurm_submit_array(results: list[SampleResult], binary: Path,
                        threads: int, max_parallel: int,
                        log_dir: Path, run_dir: Path,
                        mem_gb: int = 30, time_limit: str = "24:00:00",
                        wait: bool = True,
                        log=None) -> list[SampleResult]:
    """Materialize taiji_array.sbatch from template, submit with
    `sbatch -a 1-N%MAX_PARALLEL`, optionally wait for completion, then
    populate per-sample SampleResult fields by inspecting outputs."""
    if not SBATCH_TEMPLATE.exists():
        raise SystemExit(f"sbatch template missing: {SBATCH_TEMPLATE}")
    if shutil.which("sbatch") is None:
        raise SystemExit("[taiji-runner] --executor slurm requires `sbatch` on PATH")

    log_dir.mkdir(parents=True, exist_ok=True)
    config_list = run_dir / "taiji_config_files.txt"
    if not config_list.exists():
        raise SystemExit(f"[taiji-runner] config list missing: {config_list}")

    # Materialize the sbatch script with literal substitutions.
    job_name = f"taiji-{run_dir.name}"
    template = SBATCH_TEMPLATE.read_text()
    script_text = (template
        .replace("__JOB_NAME__",      job_name)
        .replace("__LOG_DIR__",       str(log_dir))
        .replace("__CPUS_PER_TASK__", str(threads))
        .replace("__MEM_GB__",        str(mem_gb))
        .replace("__TIME_LIMIT__",    time_limit)
        .replace("__TAIJI_BINARY__",  str(binary))
    )
    sbatch_path = run_dir / "taiji_array.sbatch"
    sbatch_path.write_text(script_text)
    sbatch_path.chmod(0o755)
    print(f"[taiji-runner] wrote {sbatch_path}", file=sys.stderr)

    # Submit.
    n = len(results)
    cap = max(1, min(max_parallel, n))
    array_spec = f"1-{n}%{cap}"
    cmd = ["sbatch", "--parsable", "-a", array_spec,
           str(sbatch_path), str(config_list)]
    print(f"[taiji-runner] $ {' '.join(cmd)}", file=sys.stderr)
    proc = subprocess.run(cmd, capture_output=True, text=True, check=False)
    if proc.returncode != 0:
        print(f"[taiji-runner] sbatch failed (exit {proc.returncode})",
              file=sys.stderr)
        print(f"  stdout: {proc.stdout.strip()}", file=sys.stderr)
        print(f"  stderr: {proc.stderr.strip()}", file=sys.stderr)
        for r in results:
            r.error = "sbatch failed"
        return results

    # `sbatch --parsable` returns just the job ID (or "<jobid>;<cluster>").
    job_id = proc.stdout.strip().split(";")[0]
    print(f"[taiji-runner] submitted SLURM array job {job_id} "
          f"({n} tasks, max {cap} concurrent)", file=sys.stderr)

    if not wait:
        # Async mode: don't block. SampleResults are unpopulated; user is
        # expected to re-run the runner with --validate-only later, or
        # poll the workflow log.
        for r in results:
            r.error = (f"submitted async (job {job_id}); rerun with "
                       "--validate-only after `squeue -j {job_id}` is empty")
        return results

    # Sync mode: poll squeue until the job is gone.
    # Correctness notes:
    #   - squeue returns exit 0 with empty stdout when no tasks remain in the
    #     queue (COMPLETED/FAILED/CANCELLED all leave the queue).  We exit only
    #     on empty stdout, never on a non-zero returncode, because a transient
    #     SLURM blip (timeout, scheduler restart) returns non-zero even when the
    #     job is still running.
    #   - COMPLETING is a brief post-run cleanup state that squeue *does* show
    #     (with non-empty stdout), so exiting on empty stdout handles it correctly
    #     — we only break after every task has left the queue entirely.
    #   - After the queue empties we call sacct for the authoritative final state,
    #     which lets us distinguish COMPLETED from FAILED/CANCELLED.
    print(f"[taiji-runner] waiting on job {job_id} (Ctrl-C is safe; SLURM keeps running)...",
          file=sys.stderr)
    poll_interval = 30
    while True:
        try:
            check = subprocess.run(
                ["squeue", "-j", job_id, "--noheader", "--format=%T %i_%a"],
                capture_output=True, text=True, check=False, timeout=15)
        except subprocess.TimeoutExpired:
            time.sleep(poll_interval)
            continue
        if not check.stdout.strip():
            # Empty output = no tasks remain in the queue.
            break
        running = len(check.stdout.strip().splitlines())
        print(f"[taiji-runner]   ... {running} task(s) still in queue",
              file=sys.stderr)
        time.sleep(poll_interval)

    # Confirm final states via sacct (authoritative; survives after queue purge).
    task_states: dict[str, str] = {}
    if shutil.which("sacct"):
        try:
            sa = subprocess.run(
                ["sacct", "-j", job_id, "--format=JobID,State", "--noheader",
                 "--parsable2"],
                capture_output=True, text=True, check=False, timeout=30)
            for line in sa.stdout.strip().splitlines():
                parts = line.split("|")
                if len(parts) >= 2:
                    task_states[parts[0].strip()] = parts[1].strip()
        except Exception:
            pass

    failed_tasks = [jid for jid, st in task_states.items()
                    if st not in ("COMPLETED", "RUNNING", "PENDING", "COMPLETING", "")]
    if failed_tasks:
        print(f"[taiji-runner] sacct reports non-COMPLETED tasks: "
              f"{', '.join(f'{j}={task_states[j]}' for j in failed_tasks)}",
              file=sys.stderr)
    print(f"[taiji-runner] job {job_id} left the queue", file=sys.stderr)

    # Validate per-sample outputs.
    for r in results:
        r.n_generanks, r.n_edges, r.n_nodes = _count_outputs(r.output_dir)
        # SLURM logs land at log/slurm-<jobid>_<taskidx>.{out,err}; capture
        # tail of stderr if present.
        # We don't know the array index per sample without parsing — keep
        # it simple by leaving stderr_log unset (samples can be matched by
        # output_dir + GeneRanks presence).
        if r.n_generanks == 0:
            r.error = ("Taiji produced no GeneRanks.tsv on SLURM; check "
                       f"{log_dir}/slurm-{job_id}_*.{{out,err}}")
            r.exit_code = 1
        else:
            r.exit_code = 0

        _log_sample(log, r)

    return results


def run_samples(results: list[SampleResult], binary: Path, threads: int,
                log_dir: Path, parallel: int = 1,
                continue_on_error: bool = False,
                log=None) -> list[SampleResult]:
    """Run each sample's taiji invocation. Sequential by default; --parallel N
    runs N samples concurrently (each with its own +RTS -N<threads>, so
    actual core usage is parallel*threads — set carefully)."""

    def _one(r: SampleResult) -> SampleResult:
        return invoke_taiji(r, binary, threads, log_dir)

    completed: list[SampleResult] = []
    if parallel <= 1:
        for r in results:
            r = _one(r)
            completed.append(r)
            _log_sample(log, r)
            if not r.ok and not continue_on_error:
                # Mark remaining samples as not run.
                for skipped in results[len(completed):]:
                    skipped.error = "skipped (prior sample failed)"
                    completed.append(skipped)
                    _log_sample(log, skipped)
                break
    else:
        with concurrent.futures.ThreadPoolExecutor(max_workers=parallel) as ex:
            futures = {ex.submit(_one, r): r for r in results}
            for fut in concurrent.futures.as_completed(futures):
                r = fut.result()
                completed.append(r)
                _log_sample(log, r)

    return completed


def _log_sample(log, result: SampleResult) -> None:
    if log is None:
        return
    try:
        log.append_custom(
            stage=f"taiji_run.{result.sample}",
            title=f"Taiji run — {result.sample}",
            payload={
                "sample":            result.sample,
                "config_path":       str(result.config_path),
                "tsv_path":          str(result.tsv_path),
                "output_dir":        str(result.output_dir),
                "exit_code":         result.exit_code,
                "duration_s":        result.duration_s,
                "n_generanks_files": result.n_generanks,
                "n_edges_files":     result.n_edges,
                "n_nodes_files":     result.n_nodes,
                "stderr_tail":       result.stderr_tail,
                "stdout_log":        str(result.stdout_log) if result.stdout_log else None,
                "stderr_log":        str(result.stderr_log) if result.stderr_log else None,
                "error":             result.error,
            },
            status="pass" if result.ok else "fail",
        )
    except Exception:
        # Logging never breaks the runner.
        pass


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _print_report(results: list[SampleResult]) -> None:
    print("\n=== taiji-runner per-sample report ===")
    width = max((len(r.sample) for r in results), default=8)
    for r in results:
        tag = "OK  " if r.ok else "FAIL"
        print(f"  [{tag}] {r.sample:<{width}}  exit={r.exit_code:>3}  "
              f"duration={r.duration_s:>6.1f}s  "
              f"generanks={r.n_generanks}  edges={r.n_edges}  nodes={r.n_nodes}")
        if r.error:
            print(f"           -> {r.error}")
    n_ok = sum(1 for r in results if r.ok)
    print(f"\nOverall: {n_ok}/{len(results)} samples ok\n")


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("--xlsx",      required=True, type=Path)
    p.add_argument("--template",  required=True, type=Path,
                   help="Per-sample formula config template (3 placeholders)")
    p.add_argument("--binary",    required=False, type=Path, default=None,
                   help="Path to the taiji binary (or use $TAIJI_BINARY)")
    p.add_argument("--run-dir",   required=True, type=Path,
                   help="Where Input/, Output/Partial/, and config list land")
    p.add_argument("--repo-root", type=Path, default=None,
                   help="For ${REPO_ROOT} substitution. Default: autodetect "
                        "from this script's grandparent's parent.")
    p.add_argument("--threads", type=int, default=4,
                   help="Threads per sample run (passed as +RTS -N<n>).")
    p.add_argument("--parallel", type=int, default=None,
                   help="Samples to run concurrently. Default: 1 for local "
                        "executor, 10 for SLURM (each sample uses <threads> cores).")
    p.add_argument("--executor", choices=["auto", "local", "slurm"],
                   default="auto",
                   help="Where to run Taiji. 'auto' picks slurm when sbatch is "
                        "on PATH, else local.")
    p.add_argument("--max-parallel", type=int, default=10,
                   help="SLURM array concurrency cap (sbatch -a 1-N%%MAX). "
                        "Default 10. Ignored under --executor local (use --parallel).")
    p.add_argument("--slurm-mem-gb", type=int, default=30,
                   help="Per-task memory request for SLURM (GB). Default 30.")
    p.add_argument("--slurm-time", default="24:00:00",
                   help="Per-task time limit for SLURM. Default 24:00:00.")
    p.add_argument("--no-wait", action="store_true",
                   help="SLURM only: submit and return immediately; don't poll squeue.")
    p.add_argument("--samples", default=None,
                   help="Comma-separated subset of sample names to run.")
    p.add_argument("--prepare-only", action="store_true",
                   help="Write per-sample TSVs + configs; do NOT invoke taiji.")
    p.add_argument("--continue-on-error", action="store_true",
                   help="Keep running remaining samples if one fails.")
    p.add_argument("--json", action="store_true",
                   help="Machine-readable output instead of text report.")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    if args.repo_root is None:
        # skills/taiji-runner/scripts/run_taiji.py -> ../../../ = workspace root
        args.repo_root = SCRIPT_DIR.parent.parent.parent
    args.repo_root = args.repo_root.resolve()
    args.run_dir = args.run_dir.resolve()
    args.run_dir.mkdir(parents=True, exist_ok=True)

    # 1. Split xlsx into per-sample SampleSpecs.
    print(f"[taiji-runner] reading {args.xlsx}", file=sys.stderr)
    samples = split_xlsx(args.xlsx)
    if not samples:
        print("ERROR: no samples found in xlsx", file=sys.stderr)
        return 2
    print(f"[taiji-runner] {len(samples)} sample(s) found: "
          f"{', '.join(s.name for s in samples)}", file=sys.stderr)
    if args.samples:
        wanted = {s.strip() for s in args.samples.split(",")}
        samples = [s for s in samples if s.name in wanted]
        if not samples:
            print(f"ERROR: --samples filter matched nothing", file=sys.stderr)
            return 2
        print(f"[taiji-runner] filtered to: "
              f"{', '.join(s.name for s in samples)}", file=sys.stderr)

    # 2. Prepare per-sample TSVs + configs + dir tree.
    results = []
    for s in samples:
        results.append(prepare_sample(
            s, args.run_dir, args.template, args.repo_root))

    # Emit the canonical config-files index (matches existing UCSD wrapper).
    config_list = args.run_dir / "taiji_config_files.txt"
    config_list.write_text(
        "\n".join(str(r.config_path) for r in results) + "\n"
    )
    print(f"[taiji-runner] wrote {config_list}", file=sys.stderr)

    if args.prepare_only:
        print("[taiji-runner] --prepare-only: stopping before taiji invocation",
              file=sys.stderr)
        if args.json:
            print(json.dumps({"prepared": [
                {"sample": r.sample,
                 "config": str(r.config_path),
                 "tsv":    str(r.tsv_path),
                 "output_dir": str(r.output_dir)} for r in results
            ]}, indent=2))
        return 0

    # 3. Resolve binary.
    binary = args.binary or os.environ.get("TAIJI_BINARY")
    if not binary:
        print("ERROR: --binary or $TAIJI_BINARY required for actual run",
              file=sys.stderr)
        return 3
    binary = Path(binary).resolve()
    if not binary.exists():
        print(f"ERROR: binary not found: {binary}", file=sys.stderr)
        return 3
    try:
        binary.chmod(binary.stat().st_mode | 0o111)
    except OSError:
        pass

    # 4. Pick executor (auto-detect SLURM if available unless overridden).
    executor = detect_executor(args.executor)
    print(f"[taiji-runner] executor: {executor}"
          + ("  (sbatch detected)" if executor == "slurm"
             and args.executor == "auto" else ""),
          file=sys.stderr)

    # Default --parallel: 1 for local (memory-conservative), 10 for slurm
    # (matches the user's ask for "max 10 in parallel").
    if args.parallel is None:
        args.parallel = 10 if executor == "slurm" else 1

    log = _attach_log(work_dir=args.run_dir)
    log_dir = args.run_dir / "log"

    if executor == "slurm":
        completed = _slurm_submit_array(
            results, binary, args.threads,
            max_parallel=args.max_parallel,
            log_dir=log_dir, run_dir=args.run_dir,
            mem_gb=args.slurm_mem_gb, time_limit=args.slurm_time,
            wait=not args.no_wait, log=log,
        )
    else:
        completed = run_samples(
            results, binary, args.threads, log_dir,
            parallel=min(args.parallel, args.max_parallel),
            continue_on_error=args.continue_on_error,
            log=log,
        )

    if args.json:
        print(json.dumps({
            "samples": [{
                "sample":            r.sample,
                "ok":                r.ok,
                "exit_code":         r.exit_code,
                "duration_s":        r.duration_s,
                "n_generanks_files": r.n_generanks,
                "n_edges_files":     r.n_edges,
                "n_nodes_files":     r.n_nodes,
                "error":             r.error,
                "config":            str(r.config_path),
                "output_dir":        str(r.output_dir),
            } for r in completed],
        }, indent=2))
    else:
        _print_report(completed)

    return 0 if all(r.ok for r in completed) else 1


if __name__ == "__main__":
    sys.exit(main())
