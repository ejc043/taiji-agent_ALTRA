"""Build the two-sheet xlsx input that bulk Taiji reads.

Usage (minimal):

    python build_taiji_input.py \
        --data-dir /path/to/processed \
        --samples  ./samples.csv \
        --genome   mm10 \
        --out      ./taiji.input.xlsx

The script is dependency-light (only openpyxl + PyYAML) and runs on Python 3.9+.
It produces two sheets:

  1. Active          — one row per assay track (RNA-seq / ATAC-seq / HiC)
  2. active_metadata — one row per group with genome FASTA + GTF

See ../references/sheet_schema.md for the full schema and rationale.
"""

from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font

# ---------------------------------------------------------------------------
# Constants — the exact schema bulk Taiji expects
# ---------------------------------------------------------------------------

ACTIVE_COLUMNS = ("type", "id", "group", "rep", "path", "tags", "format", "cohort")
METADATA_COLUMNS = ("Submitter_ID", "Case_ID", "vcf_Location", "gtf_Location")

TYPE_RNA = "RNA-seq"
TYPE_ATAC = "ATAC-seq"
TYPE_HIC = "HiC"

# tag/format conventions per track type
TRACK_SPEC = {
    TYPE_RNA:  {"id_prefix": "RNA-",  "tags": "GeneQuant",      "format": None},
    TYPE_ATAC: {"id_prefix": "ATAC-", "tags": None,             "format": "NarrowPeak"},
    TYPE_HIC:  {"id_prefix": "HiC_",  "tags": "ChromosomeLoop", "format": None},
}


def _load_detector():
    """Try to import detect_dataset_type from the sibling detect-dataset-type skill.

    Returns the function, or None if the sibling skill isn't present. Kept as a
    soft dependency so this script still runs in a standalone install of just
    build-taiji-input.
    """
    sibling = (
        Path(__file__).resolve().parent.parent.parent
        / "detect-dataset-type"
        / "scripts"
    )
    if sibling.exists() and str(sibling) not in sys.path:
        sys.path.insert(0, str(sibling))
    try:
        from detect import detect_dataset_type  # type: ignore[import-not-found]
        return detect_dataset_type
    except ImportError:
        return None


def _find_epitensor(genome_name: str) -> Path | None:
    """Return the vendored epitensor loops file for genome_name, or None.

    Looks for exactly one file under <repo_root>/epitensor/<genome_name>/.
    Returns None silently when the genome has no vendored file.
    """
    epitensor_dir = (
        Path(__file__).resolve().parent.parent.parent.parent
        / "epitensor" / genome_name
    )
    if not epitensor_dir.is_dir():
        return None
    files = [f for f in epitensor_dir.iterdir() if f.is_file()]
    return files[0] if files else None


def _attach_log():
    """Soft-attach to the workflow-log skill's active run. Returns None if the
    skill isn't installed alongside or no run is active. Never raises."""
    try:
        sibling = (
            Path(__file__).resolve().parent.parent.parent
            / "workflow-log" / "scripts"
        )
        if sibling.exists() and str(sibling) not in sys.path:
            sys.path.insert(0, str(sibling))
        from log import TaijiLog  # type: ignore[import-not-found]
        return TaijiLog.attach_active(work_dir=Path.cwd())
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclass
class SampleRow:
    group: str
    cohort: str
    rep: int = 1
    hic_path: Path | None = None  # per-sample HiC override


@dataclass
class GenomeEntry:
    name: str
    fasta: Path
    gtf: Path
    hic: Path | None

    @property
    def has_todo_paths(self) -> bool:
        # A heuristic: any path containing a literal "TODO" marker or a
        # clearly-placeholder "/path/to/" prefix is treated as unset.
        for p in (self.fasta, self.gtf, self.hic):
            if p is None:
                continue
            s = str(p)
            if "TODO" in s or s.startswith("/path/to/"):
                return True
        return False


# ---------------------------------------------------------------------------
# I/O helpers
# ---------------------------------------------------------------------------


def load_genomes(config_path: Path) -> dict[str, GenomeEntry]:
    with config_path.open() as fh:
        raw = yaml.safe_load(fh) or {}
    out: dict[str, GenomeEntry] = {}
    for name, entry in raw.items():
        out[name] = GenomeEntry(
            name=name,
            fasta=Path(entry["fasta"]),
            gtf=Path(entry["gtf"]),
            hic=Path(entry["hic"]) if entry.get("hic") else None,
        )
    return out


def load_samples(csv_path: Path) -> list[SampleRow]:
    rows: list[SampleRow] = []
    delimiter = "\t" if csv_path.suffix.lower() in (".tsv", ".tab") else ","
    with csv_path.open(newline="") as fh:
        reader = csv.DictReader(fh, delimiter=delimiter)
        missing = {"group", "cohort"} - set(reader.fieldnames or ())
        if missing:
            raise SystemExit(
                f"samples CSV {csv_path} is missing required columns: {sorted(missing)}"
            )
        for i, row in enumerate(reader, start=2):
            group = (row.get("group") or "").strip()
            cohort = (row.get("cohort") or "").strip()
            if not group or not cohort:
                raise SystemExit(
                    f"{csv_path}: row {i} has empty group or cohort"
                )
            rep = int(row.get("rep") or 1)
            hic_raw = (row.get("hic_path") or "").strip()
            hic = Path(hic_raw) if hic_raw else None
            rows.append(SampleRow(group=group, cohort=cohort, rep=rep, hic_path=hic))
    return rows


# ---------------------------------------------------------------------------
# Path resolution
# ---------------------------------------------------------------------------


def resolve_path(data_dir: Path, pattern: str, group: str) -> Path:
    """Interpolate `{group}` in the pattern and join against data_dir."""
    return data_dir / pattern.format(group=group)


# ---------------------------------------------------------------------------
# Row assembly
# ---------------------------------------------------------------------------


def _active_row(track_type: str, group: str, rep: int, path: Path, cohort: str) -> dict:
    spec = TRACK_SPEC[track_type]
    return {
        "type": track_type,
        "id": f"{spec['id_prefix']}{group}",
        "group": group,
        # Stored as text to match the reference sheet format exactly; Taiji
        # tolerates either, but string rep keeps diffs clean against hand-edited
        # sheets where rep is quoted.
        "rep": str(rep),
        "path": str(path),
        # `None` rather than "" so the cell is genuinely empty in the xlsx,
        # matching how Taiji's reference sheets are authored.
        "tags": spec["tags"],
        "format": spec["format"],
        "cohort": cohort,
    }


def assemble_active_rows(
    samples: Iterable[SampleRow],
    data_dir: Path,
    rna_pattern: str,
    atac_pattern: str,
    hic_pattern: str | None,
    genome: GenomeEntry,
    include_hic: bool,
    strict: bool,
) -> tuple[list[dict], list[SampleRow], list[str], list[str]]:
    """Return (rows, kept_samples, errors, warnings).

    `kept_samples` is the subset of `samples` that ended up in the sheet — in
    non-strict mode, groups with any missing required modality are dropped
    entirely so the output stays self-consistent (no orphaned RNA rows for
    samples missing ATAC, etc.). In strict mode, nothing is dropped and the
    errors are returned for the caller to handle.
    """
    rows: list[dict] = []
    kept: list[SampleRow] = []
    errors: list[str] = []
    warnings: list[str] = []
    seen_ids: set[str] = set()

    for s in samples:
        rna_path = resolve_path(data_dir, rna_pattern, s.group)
        atac_path = resolve_path(data_dir, atac_pattern, s.group)

        missing: list[str] = []
        if not rna_path.exists():
            missing.append(f"RNA-seq at {rna_path}")
        if not atac_path.exists():
            missing.append(f"ATAC-seq at {atac_path}")

        if missing:
            msg = f"group '{s.group}': missing {', '.join(missing)}"
            if strict:
                errors.append(msg)
                # In strict, still emit the rows we have so errors are visible
                # per-row in the eventual sheet (if the user overrides).
                if rna_path.exists():
                    rows.append(_active_row(TYPE_RNA, s.group, s.rep, rna_path, s.cohort))
                if atac_path.exists():
                    rows.append(_active_row(TYPE_ATAC, s.group, s.rep, atac_path, s.cohort))
            else:
                warnings.append(msg + " — dropping group from output")
                continue  # drop the whole sample, don't emit any rows or keep it
        else:
            rows.append(_active_row(TYPE_RNA, s.group, s.rep, rna_path, s.cohort))
            rows.append(_active_row(TYPE_ATAC, s.group, s.rep, atac_path, s.cohort))

        # Only reach here if: strict (errors tolerated), or all required files present.
        kept.append(s)

        # HiC: per-sample override > discovered file > genome default > omit
        if include_hic:
            hic_path: Path | None = None
            if s.hic_path is not None:
                hic_path = s.hic_path
            elif hic_pattern is not None:
                candidate = resolve_path(data_dir, hic_pattern, s.group)
                if candidate.exists():
                    hic_path = candidate
            if hic_path is None and genome.hic is not None:
                hic_path = genome.hic
            # Epitensor fallback: vendored per-genome enhancer loops shipped with
            # the repo. Used when no HiC is supplied or configured in genomes.yml.
            if hic_path is None:
                epitensor = _find_epitensor(genome.name)
                if epitensor is not None:
                    hic_path = epitensor
                    warnings.append(
                        f"group '{s.group}': no HiC supplied; "
                        f"using vendored epitensor fallback ({epitensor.name})"
                    )
            if hic_path is not None:
                if not hic_path.exists():
                    warnings.append(
                        f"group '{s.group}': HiC file declared but not found at {hic_path}"
                    )
                rows.append(_active_row(TYPE_HIC, s.group, s.rep, hic_path, s.cohort))
            else:
                warnings.append(f"group '{s.group}': no HiC file resolved (skipping HiC row)")

    # uniqueness check
    for r in rows:
        if r["id"] in seen_ids:
            errors.append(f"duplicate id in Active sheet: {r['id']}")
        seen_ids.add(r["id"])

    return rows, kept, errors, warnings


def assemble_metadata_rows(samples: Iterable[SampleRow], genome: GenomeEntry) -> list[dict]:
    return [
        {
            "Submitter_ID": s.group,
            "Case_ID": s.group,
            "vcf_Location": str(genome.fasta),
            "gtf_Location": str(genome.gtf),
        }
        for s in samples
    ]


# ---------------------------------------------------------------------------
# Cross-sheet validation
# ---------------------------------------------------------------------------


def validate_cross_sheet(active_rows: list[dict], metadata_rows: list[dict]) -> list[str]:
    errors: list[str] = []
    groups_in_active = {r["group"] for r in active_rows}
    groups_in_meta = {r["Submitter_ID"] for r in metadata_rows}

    only_active = groups_in_active - groups_in_meta
    only_meta = groups_in_meta - groups_in_active
    if only_active:
        errors.append(f"groups in Active missing from active_metadata: {sorted(only_active)}")
    if only_meta:
        errors.append(f"groups in active_metadata never referenced in Active: {sorted(only_meta)}")

    # every group needs at least one RNA + one ATAC
    by_group: dict[str, set[str]] = defaultdict(set)
    for r in active_rows:
        by_group[r["group"]].add(r["type"])
    for g, types in by_group.items():
        if TYPE_RNA not in types:
            errors.append(f"group '{g}': no RNA-seq row in Active")
        if TYPE_ATAC not in types:
            errors.append(f"group '{g}': no ATAC-seq row in Active")

    # cohort consistency within group
    cohorts: dict[str, set[str]] = defaultdict(set)
    for r in active_rows:
        cohorts[r["group"]].add(r["cohort"])
    for g, cs in cohorts.items():
        if len(cs) > 1:
            errors.append(f"group '{g}': inconsistent cohort labels {sorted(cs)}")

    return errors


# ---------------------------------------------------------------------------
# xlsx writer
# ---------------------------------------------------------------------------


def write_xlsx(active_rows: list[dict], metadata_rows: list[dict], out: Path) -> None:
    wb = Workbook()
    # Active
    ws1 = wb.active
    ws1.title = "Active"
    ws1.append(list(ACTIVE_COLUMNS))
    for col_idx, _ in enumerate(ACTIVE_COLUMNS, start=1):
        ws1.cell(row=1, column=col_idx).font = Font(bold=True)
    for r in active_rows:
        ws1.append([r[c] for c in ACTIVE_COLUMNS])

    # active_metadata
    ws2 = wb.create_sheet("active_metadata")
    ws2.append(list(METADATA_COLUMNS))
    for col_idx, _ in enumerate(METADATA_COLUMNS, start=1):
        ws2.cell(row=1, column=col_idx).font = Font(bold=True)
    # unique groups, preserving input order
    seen = set()
    for m in metadata_rows:
        key = m["Submitter_ID"]
        if key in seen:
            continue
        seen.add(key)
        ws2.append([m[c] for c in METADATA_COLUMNS])

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------


def print_summary(active_rows: list[dict], metadata_rows: list[dict], out: Path) -> None:
    groups = {r["group"] for r in active_rows}
    type_counts = Counter(r["type"] for r in active_rows)
    cohort_counts = Counter(r["cohort"] for r in active_rows if r["type"] == TYPE_RNA)
    print(f"\nWrote {out}")
    print(f"  groups        : {len(groups)}")
    print(f"  RNA-seq rows  : {type_counts.get(TYPE_RNA, 0)}")
    print(f"  ATAC-seq rows : {type_counts.get(TYPE_ATAC, 0)}")
    print(f"  HiC rows      : {type_counts.get(TYPE_HIC, 0)}")
    print(f"  metadata rows : {len({m['Submitter_ID'] for m in metadata_rows})}")
    if cohort_counts:
        breakdown = ", ".join(f"{k}={v}" for k, v in sorted(cohort_counts.items()))
        print(f"  cohort split  : {breakdown}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Build the bulk Taiji input xlsx (Active + active_metadata sheets).",
    )
    p.add_argument("--data-dir", required=True, type=Path,
                   help="Directory containing processed per-sample files")
    p.add_argument("--samples", required=True, type=Path,
                   help="CSV with at least columns: group, cohort (optional: rep, hic_path)")
    p.add_argument("--genome", required=True,
                   help="Genome build name, must exist in the genomes config (e.g. mm10, hg38)")
    p.add_argument("--out", required=True, type=Path,
                   help="Output xlsx path")
    p.add_argument("--rna-pattern", default="{group}.rna.txt",
                   help="Filename pattern for RNA-seq files under --data-dir (default: %(default)s)")
    p.add_argument("--atac-pattern", default="{group}.narrowPeak",
                   help="Filename pattern for ATAC-seq files under --data-dir (default: %(default)s)")
    p.add_argument("--hic-pattern", default=None,
                   help="Filename pattern for per-sample HiC files; if unset, the per-genome "
                        "default from the genomes config is used")
    p.add_argument("--no-hic", action="store_true",
                   help="Skip HiC rows entirely even if files are available")
    p.add_argument("--genome-config", type=Path, default=None,
                   help="Path to an alternate genomes YAML (default: assets/genomes.yml next to this script)")
    strictness = p.add_mutually_exclusive_group()
    strictness.add_argument("--strict", dest="strict", action="store_true", default=True,
                            help="Fail on any missing RNA or ATAC file (default)")
    strictness.add_argument("--no-strict", dest="strict", action="store_false",
                            help="Warn and skip missing files instead of failing")
    p.add_argument("--skip-data-type-check", action="store_true",
                   help="Skip the bulk/single-cell pre-flight gate. Default is to refuse "
                        "running on single-cell or mixed datasets.")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    # resolve genome config
    genome_config = args.genome_config
    if genome_config is None:
        genome_config = Path(__file__).resolve().parent.parent / "assets" / "genomes.yml"
    if not genome_config.exists():
        print(f"ERROR: genome config not found: {genome_config}", file=sys.stderr)
        return 2

    genomes = load_genomes(genome_config)
    if args.genome not in genomes:
        print(
            f"ERROR: genome '{args.genome}' not in {genome_config}. "
            f"Available: {sorted(genomes)}",
            file=sys.stderr,
        )
        return 2
    genome = genomes[args.genome]

    if genome.has_todo_paths:
        print(
            f"WARNING: genome '{args.genome}' entry in {genome_config} still has placeholder "
            f"paths. Edit this file before running Taiji on the output.",
            file=sys.stderr,
        )

    if not args.data_dir.exists():
        print(f"ERROR: --data-dir does not exist: {args.data_dir}", file=sys.stderr)
        return 2

    # Pre-flight: refuse to proceed on single-cell or mixed datasets. Soft-skip
    # if the sibling detect-dataset-type skill isn't installed alongside this
    # one (standalone installs of just build-taiji-input should still work).
    if not args.skip_data_type_check:
        detector = _load_detector()
        if detector is None:
            print(
                "NOTE: detect-dataset-type skill not found; skipping bulk/SC pre-flight. "
                "Pass --skip-data-type-check to silence this notice.",
                file=sys.stderr,
            )
        else:
            # Use non-recursive scan: pseudobulk output directories have
            # _peaks.rds files inside atac/rds/ that would falsely trigger
            # the mixed-dataset gate when scanning recursively. The actual
            # bulk files (.tsv, .narrowPeak) live in subdirectories anyway,
            # so the top-level scan still catches stray SC objects.
            detect_result = detector([args.data_dir], recursive=False, strict_mixed=True)
            if detect_result.classification == "single-cell":
                modality = getattr(detect_result, "sc_modality", None)
                modality_note = {
                    "multiome": (
                        " RNA and ATAC are paired at the cell level (multiome); "
                        "route to a single-cell Taiji workflow that accepts paired "
                        "inputs (e.g. sc-Taiji with AnnData + fragments)."
                    ),
                    "separate-assay": (
                        " RNA and ATAC come from DIFFERENT cell populations. "
                        "Co-embed them first via Signac's integrate_atac workflow "
                        "(reciprocal-LSI anchors + IntegrateEmbeddings): "
                        "https://stuartlab.org/signac/articles/integrate_atac . "
                        "Afterwards, route to a single-cell Taiji workflow."
                    ),
                    "sc-undetermined": (
                        " The RNA/ATAC modality layout could not be inferred from "
                        "filenames. Confirm whether this is multiome (same cells) "
                        "or separate-assay (different cells, needs co-embedding "
                        "via https://stuartlab.org/signac/articles/integrate_atac) "
                        "before choosing a downstream workflow."
                    ),
                }.get(modality, "")
                print(
                    f"ERROR: {args.data_dir} contains single-cell data "
                    f"({sorted(detect_result.sc_files.keys())}"
                    f"{f'; sc_modality={modality}' if modality else ''}). "
                    "build-taiji-input only supports bulk RNA-seq + ATAC-seq "
                    f"(+ optional HiC).{modality_note} Override with "
                    "--skip-data-type-check if you know what you're doing.",
                    file=sys.stderr,
                )
                return 2
            if detect_result.classification == "mixed":
                print(
                    f"ERROR: {args.data_dir} contains both bulk and single-cell "
                    "signatures. Bulk Taiji cannot consume mixed datasets. Split "
                    "the directories or remove the minority files, then re-run.",
                    file=sys.stderr,
                )
                for e in detect_result.errors:
                    print(f"  - {e}", file=sys.stderr)
                return 2
            if detect_result.classification == "unknown":
                print(
                    f"WARNING: no bulk signatures found in {args.data_dir}. "
                    "Proceeding because file-pattern overrides may apply; "
                    f"top extensions seen: "
                    f"{sorted(detect_result.other_exts.items(), key=lambda kv: -kv[1])[:5]}",
                    file=sys.stderr,
                )

    samples = load_samples(args.samples)
    if not samples:
        print("ERROR: samples CSV has no rows", file=sys.stderr)
        return 2

    active_rows, kept_samples, errors, warnings = assemble_active_rows(
        samples=samples,
        data_dir=args.data_dir,
        rna_pattern=args.rna_pattern,
        atac_pattern=args.atac_pattern,
        hic_pattern=args.hic_pattern,
        genome=genome,
        include_hic=not args.no_hic,
        strict=args.strict,
    )
    # Metadata rows match the groups that actually made it into Active; in
    # non-strict mode this is the pruned set, in strict mode it's all samples.
    metadata_rows = assemble_metadata_rows(kept_samples if not args.strict else samples, genome)
    errors.extend(validate_cross_sheet(active_rows, metadata_rows))

    for w in warnings:
        print(f"WARN: {w}", file=sys.stderr)

    if errors:
        print("\nERRORS:", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        print(f"\n{len(errors)} error(s); refusing to write {args.out}", file=sys.stderr)
        return 1

    write_xlsx(active_rows, metadata_rows, args.out)
    print_summary(active_rows, metadata_rows, args.out)

    # Best-effort log entry. Pull gate fields off the local detect_result if it
    # exists (it does unless --skip-data-type-check or the sibling skill was
    # missing). All fields are optional from the log's perspective.
    log = _attach_log()
    if log is not None:
        try:
            gate_cls = None
            gate_mod = None
            try:
                gate_cls = detect_result.classification           # type: ignore[name-defined]
                gate_mod = getattr(detect_result, "sc_modality", None)  # type: ignore[name-defined]
            except NameError:
                pass
            dropped_names = [s.name for s in samples
                             if s.name not in {k.name for k in kept_samples}]
            log.append_build_input({
                "xlsx_path":          str(args.out),
                "n_active_rows":      len(active_rows),
                "n_dropped_samples":  len(dropped_names),
                "dropped_reasons":    {n: "missing files / failed cross-sheet validation"
                                       for n in dropped_names},
                "samples_path":       str(args.samples),
                "data_dir":           str(args.data_dir),
                "genome":             genome,
                "gate_classification": gate_cls,
                "gate_sc_modality":    gate_mod,
                "cohorts":             sorted({r.get("group", r.get("cohort", ""))
                                               for r in metadata_rows
                                               if isinstance(r, dict)}),
                "warnings":            warnings,
                "status":              "pass",
            })
        except Exception:
            pass

    return 0


if __name__ == "__main__":
    sys.exit(main())
