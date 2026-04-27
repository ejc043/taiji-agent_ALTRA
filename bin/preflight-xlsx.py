"""preflight-xlsx — verify every file referenced in a Taiji input xlsx exists.

Reads the `Active` sheet, checks that every path is resolvable + non-empty.
Reads the `active_metadata` sheet, checks the genome / annotation paths.
Exits non-zero with a clean error report if anything's missing — saves you
the cost of a failed Taiji run that bombs 30 minutes in on a bad path.

Usage:
    python bin/preflight-xlsx.py runs/<name>/taiji_input.xlsx
    python bin/preflight-xlsx.py runs/<name>/taiji_input.xlsx --json
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("preflight: openpyxl required (`micromamba install openpyxl` or "
          "`pip install openpyxl`)", file=sys.stderr)
    sys.exit(2)


def _check_path(p: str | None, kind: str) -> dict:
    rec = {"kind": kind, "path": p, "ok": False, "error": None,
           "size_bytes": 0}
    if not p:
        rec["error"] = "empty path"
        return rec
    pp = Path(p)
    if not pp.exists():
        rec["error"] = "does not exist"
        return rec
    if not pp.is_file():
        rec["error"] = "is not a regular file"
        return rec
    try:
        size = pp.stat().st_size
        rec["size_bytes"] = size
        if size == 0:
            rec["error"] = "empty file"
            return rec
    except OSError as e:
        rec["error"] = f"stat failed: {e}"
        return rec
    rec["ok"] = True
    return rec


def preflight(xlsx_path: Path) -> dict:
    if not xlsx_path.exists():
        return {"xlsx": str(xlsx_path), "ok": False,
                "error": "xlsx file not found", "checks": []}
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    checks: list[dict] = []

    # Active sheet: data files referenced in the `path` column.
    active = wb["Active"] if "Active" in wb.sheetnames else None
    if active is None:
        return {"xlsx": str(xlsx_path), "ok": False,
                "error": "Active sheet missing", "checks": []}
    headers = [c.value for c in next(active.iter_rows(min_row=1, max_row=1))]
    if "path" not in headers:
        return {"xlsx": str(xlsx_path), "ok": False,
                "error": "Active sheet has no 'path' column", "checks": []}
    path_idx = headers.index("path")
    type_idx = headers.index("type") if "type" in headers else None
    id_idx = headers.index("id") if "id" in headers else None

    for row in active.iter_rows(min_row=2, values_only=True):
        if not row or all(c is None for c in row):
            continue
        sample_id = row[id_idx] if id_idx is not None else "?"
        type_ = row[type_idx] if type_idx is not None else "?"
        kind = f"Active[{type_}/{sample_id}]"
        checks.append(_check_path(row[path_idx], kind))

    # active_metadata: vcf_Location (FASTA) + gtf_Location columns.
    md = wb["active_metadata"] if "active_metadata" in wb.sheetnames else None
    if md is not None:
        md_headers = [c.value for c in next(md.iter_rows(min_row=1, max_row=1))]
        for col_name, kind_label in (("vcf_Location", "metadata.fasta"),
                                     ("gtf_Location", "metadata.gtf")):
            if col_name not in md_headers:
                continue
            col_idx = md_headers.index(col_name)
            seen_paths: set[str] = set()
            for row in md.iter_rows(min_row=2, values_only=True):
                if not row or row[col_idx] is None:
                    continue
                p = str(row[col_idx])
                # Don't re-check the same path 1000 times if every sample
                # uses the same reference (the common case).
                if p in seen_paths:
                    continue
                seen_paths.add(p)
                checks.append(_check_path(p, kind_label))

    failures = [c for c in checks if not c["ok"]]
    return {"xlsx": str(xlsx_path), "ok": not failures, "checks": checks,
            "n_checks": len(checks), "n_failures": len(failures)}


def _human_bytes(n: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024:
            return f"{n:.1f} {unit}"
        n = n / 1024.0
    return f"{n:.1f} TB"


def _print_report(result: dict) -> None:
    print(f"=== preflight: {result['xlsx']} ===")
    if "error" in result and result.get("error"):
        print(f"  FATAL: {result['error']}")
        return
    for c in result["checks"]:
        tag = "OK   " if c["ok"] else "MISS "
        size = _human_bytes(c["size_bytes"]) if c.get("size_bytes") else "—"
        print(f"  [{tag}] {c['kind']:32s}  ({size:>9s})  {c['path']}")
        if c.get("error"):
            print(f"           -> {c['error']}")
    summary = f"{result['n_checks'] - result['n_failures']}/{result['n_checks']} OK"
    print(f"\nSummary: {summary} — {'PASS' if result['ok'] else 'FAIL'}")


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("xlsx", type=Path)
    p.add_argument("--json", action="store_true",
                   help="Machine-readable output instead of text.")
    args = p.parse_args(argv)
    result = preflight(args.xlsx)
    if args.json:
        print(json.dumps(result, indent=2))
    else:
        _print_report(result)
    return 0 if result["ok"] else 1


if __name__ == "__main__":
    sys.exit(main())
